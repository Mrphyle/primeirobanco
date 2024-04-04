import requests
import openpyxl

# Função para baixar o arquivo Excel do SharePoint Online
def baixar_excel_sharepoint(file_url):
    response = requests.get(file_url)
    if response.status_code == 200:
        with open("registros_elevadores.xlsx", "wb") as file:
            file.write(response.content)
        print("Arquivo Excel baixado com sucesso!")
    else:
        print("Erro ao baixar o arquivo Excel:", response.status_code)

# Função para calcular média e porcentagem
def calcular_media_porcentagem(registros, chave):
    total_usos = len(registros)
    contagem = {valor: registros.count(valor) for valor in set(registros)}
    media = total_usos / len(contagem)
    porcentagem = {valor: (contagem[valor] / total_usos) * 100 for valor in contagem}
    mais_usado = max(contagem, key=contagem.get)
    return media, porcentagem, mais_usado

# Função para adicionar dados ao arquivo Excel
def adicionar_dados_excel(registros, nome_arquivo):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Elevador'
    sheet['B1'] = 'Período'

    for i, registro in enumerate(registros, start=2):
        sheet[f'A{i}'] = registro['elevador']
        sheet[f'B{i}'] = registro['periodo']

    workbook.save(nome_arquivo)

# URL do arquivo Excel no SharePoint Online
file_url = "https://senacspedu-my.sharepoint.com/:x:/g/personal/diego_sgoncalves10_senacsp_edu_br/ESUPTjxSWhFHtEKgfCEFIjkBpgZ1tBgBcsf8jJQpsJbkKA"

# Baixar o arquivo Excel do SharePoint Online
baixar_excel_sharepoint(file_url)

# Abrir o arquivo Excel baixado
nome_arquivo = "registros_elevadores.xlsx"
workbook = openpyxl.load_workbook(nome_arquivo)
sheet = workbook.active

# Ler os registros do arquivo Excel
registros = [{'elevador': sheet[f'A{i}'].value, 'periodo': sheet[f'B{i}'].value} for i in range(2, sheet.max_row + 1)]

# Calcular média e porcentagem de elevadores e períodos
media_elevador, porcentagem_elevador, elevador_mais_usado = calcular_media_porcentagem([registro['elevador'] for registro in registros], 'elevador')
media_horario, porcentagem_horario, horario_mais_usado = calcular_media_porcentagem([registro['periodo'] for registro in registros], 'periodo')

# Imprimir resultados
print(f"A média de elevador usado foi {media_elevador:.2f} e o elevador mais usado é o {elevador_mais_usado} com {porcentagem_elevador[elevador_mais_usado]:.2f}%")
print(f"A média do horário usado foi {media_horario:.2f} e o horário mais usado é o {horario_mais_usado} com {porcentagem_horario[horario_mais_usado]:.2f}%")